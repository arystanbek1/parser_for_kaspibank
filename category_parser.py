from selenium import webdriver
from bs4 import BeautifulSoup
import openpyxl
import time
from main import choose_category


def parser():
    wb = openpyxl.Workbook()
    wb.create_sheet(title='Sheet1', index=0)
    sheet = wb['Sheet1']

    i = 1
    j = 1
    cell = sheet.cell(row=j, column=1)
    cell.value = 'ID'
    cell = sheet.cell(row=j, column=2)
    cell.value = 'Product_ID'
    cell = sheet.cell(row=j, column=3)
    cell.value = 'NAME'
    cell = sheet.cell(row=j, column=4)
    cell.value = 'PRICE'
    cell = sheet.cell(row=j, column=5)
    cell.value = 'DETAILS'
    cell = sheet.cell(row=j, column=6)
    cell.value = 'LINK'

    chromedriver = '/Users/arystanbekabdrahmanov/PycharmProjects/parser_123/chromedriver'
    options = webdriver.ChromeOptions()
    options.add_argument('headless')  # для открытия headless-браузера
    browser = webdriver.Chrome(executable_path=chromedriver, options=options)

    url = ''
    if choose_category == 'Айфоны':
        url = 'https://kaspi.kz/shop/c/smartphones/?q=%3AproductClass%3AApple+iPhone&page='
    elif choose_category == 'Шины':
        url = 'https://kaspi.kz/shop/c/passenger%20car%20tires/' \
              '?q=%3Acategory%3APassenger+car+tires%3Acategory%3ATires&page='
    elif choose_category == 'Морозилки':
        url = 'https://kaspi.kz/shop/c/freezers/?page='
    elif choose_category == 'Наушники':
        url = 'https://kaspi.kz/shop/c/headphones/?page='
    elif choose_category == 'Ноутбуки офисные':
        url = 'https://kaspi.kz/shop/c/notebooks/?q=%3AproductClass%3AДля+работы+и+учёбы&page='
    elif choose_category == 'Ноутбуки игровые':
        url = 'https://kaspi.kz/shop/c/notebooks/?q=%3AproductClass%3AИгровые&page='
    elif choose_category == 'Кондиционеры':
        url = 'https://kaspi.kz/shop/c/air%20conditioners/?q=%3AproductClass%3AВ+комнату&page='
    elif choose_category == 'мониторы':
        url = 'https://kaspi.kz/shop/c/monitors/?page='
    else:
        print('Не тупи, вводи из предложенного)))')

    url2 = url
    next_page_check = False
    while not next_page_check:
        url = url2 + str(i)

        browser.get(url)
        html = browser.page_source

        soup = BeautifulSoup(html, 'lxml')
        divs = soup.find(
            'div',
            class_="item-cards-grid__cards").find_all('div', class_="item-card ddl_product ddl_product_link undefined")

        try:
            browser.get(url)
            html = browser.page_source

            soup = BeautifulSoup(html, 'lxml')
            divs = soup.find(
                'div',
                class_="item-cards-grid__cards").find_all('div', class_="item-card ddl_product ddl_product_link undefined")

        except:
            print('--------- next page not found')
            break

        for d in divs:
            price = d.find('span', class_="item-card__prices-price").text.replace(' ', '').replace('₸', '')
            name = d.find('a', class_="item-card__name-link").text
            product_id = d.get('data-product-id')
            link = d.find('a', class_='item-card__name-link').get('href')

            # отдельный реквест для того, чтобы перейти на страницу продукта #####
            time.sleep(1)

            r1 = browser.get(link)
            product_html = browser.page_source

            soup_info = BeautifulSoup(product_html, 'lxml')
            try:
                details = soup_info('div', class_='item__description-text')
                details = details[0].text

            except:
                details = 'no descriptions'

            j += 1
            cell = sheet.cell(row=j, column=1)
            cell.value = j - 1
            cell = sheet.cell(row=j, column=2)
            cell.value = product_id
            cell = sheet.cell(row=j, column=3)
            cell.value = name
            cell = sheet.cell(row=j, column=4)
            cell.value = price
            cell = sheet.cell(row=j, column=5)
            cell.value = details
            cell = sheet.cell(row=j, column=6)
            cell.value = link
            print(j - 1, name,'|', price, 'тенге |||| ===========')
            file_name = 'prices_from_kaspi_kz.xlsx'
            wb.save(file_name)
        if i % 2 == 0:
            time.sleep(1)
        next_page_check = soup.findAll('li', class_="pagination__el _disabled")
        if next_page_check and next_page_check[0].contents[0] == 'Следующая →':
            break
        else:
            i += 1
            next_page_check = False

    print('PARSER TASK FINISHED NORMAL')
