import os

import telebot
import mysql.connector
from selenium import webdriver
from bs4 import BeautifulSoup
import openpyxl
import time
from smtplib_1 import send_mail

token = "5579236927:AAFulGkMYycjTMd028fBq-EIDc1H9TbGjHs"
bot = telebot.TeleBot(token)


mydb = mysql.connector.connect(
    host="localhost",
    port=3306,
    user="root",
    passwd="",
    database="parser"
)
myqursor = mydb.cursor()

registrations_variable = 0
name_mysql = 0
surname_mysql = 0
mail_mysql = 0
city_mysql = 0
number_mysql = 0
user_id = 0
choose_category = 0
choose_variable = 0
choose_for_menu = 0
parser_variable = 0


@bot.message_handler(commands=["start"])
def handle_start(message):
    user_markup = telebot.types.ReplyKeyboardMarkup(True, False)
    user_markup.row("Зарегистрироваться 📋")
    user_markup.row("Войти ️➡️")
    user_markup.row("О нас 👩‍💻")
    user_markup.row("Выйти")
    bot.send_message(message.chat.id, "Добро пожаловать в Парсинг центр 😄", reply_markup=user_markup)


def menu(message):
    user_markup = telebot.types.ReplyKeyboardMarkup(True, False)
    user_markup.row("Что вы хотите запарсить?")
    user_markup.row("Айфоны", "Шины")
    user_markup.row("Морозилки", "Наушники")
    user_markup.row("Ноутбуки офисные", "Ноутбуки игровые")
    user_markup.row("Кондиционеры", "мониторы")
    bot.send_message(message.chat.id, "Вы успешно зарегистрированы ✅", reply_markup=user_markup)


@bot.message_handler(commands=["stop"])
def handle_stop(message):
    hide_markup = telebot.types.ReplyKeyboardRemove()
    bot.send_message(message.chat.id, "пока!!!", reply_markup=hide_markup)


@bot.message_handler(content_types=["text"])     #Вторая менюшка
def handle_menu(message):
    global registrations_variable
    global user_id
    global name_mysql
    global surname_mysql
    global city_mysql
    global number_mysql
    global mail_mysql
    global choose_variable
    global choose_category
    global choose_for_menu

    if message.text == "Зарегистрироваться 📋":
        bot.send_message(message.chat.id, 'Ваше имя?')
        registrations_variable = 1
        user_id = message.from_user.id

    elif registrations_variable == 1:
        registrations_variable = 2
        bot.send_message(message.chat.id, 'Ваше фамилия?')
        name_mysql = message.text

    elif registrations_variable == 2:
        registrations_variable = 3
        bot.send_message(message.chat.id, 'Ваша почта?')
        surname_mysql = message.text

    elif registrations_variable == 3:
        registrations_variable = 4
        bot.send_message(message.chat.id, 'С какого вы города?')
        mail_mysql = message.text

    elif registrations_variable == 4:
        registrations_variable = 5
        bot.send_message(message.chat.id, 'Ваш номер телефона?')
        city_mysql = message.text

    elif registrations_variable == 5:
        registrations_variable = 6
        number_mysql = message.text
        menu(message)

    elif registrations_variable == 6:
        choose_category = message.text
        registration()
        parser1()
        send_mail(mail_mysql)
        print(mail_mysql)

    elif message.text == "Войти ️➡️":
        join(message)

    elif message.text == "Выйти":
        handle_stop(message)


def registration():
    global user_id
    global name_mysql
    global surname_mysql
    global city_mysql
    global number_mysql
    global mail_mysql
    global user_id

    sql = "insert into users(user_id,name,surname,mail,city,number,category) values(%s,%s,%s,%s,%s,%s,%s)"
    values = (user_id, name_mysql, surname_mysql, mail_mysql, city_mysql, number_mysql, choose_category)
    myqursor.execute(sql, values)
    mydb.commit()


def parser1():
    global parser_variable
    wb = openpyxl.Workbook()
    wb.create_sheet(title='Sheet1', index=0)
    sheet = wb['Sheet1']

    i = 1
    j = 1
    cell = sheet.cell(row=j, column=1)
    cell.value = 'ID'
    cell = sheet.cell(row=j, column=2)
    cell.value = 'Product_ID'
    cell = sheet.cell(row=j, column=3)
    cell.value = 'NAME'
    cell = sheet.cell(row=j, column=4)
    cell.value = 'PRICE'
    cell = sheet.cell(row=j, column=5)
    cell.value = 'DETAILS'
    cell = sheet.cell(row=j, column=6)
    cell.value = 'LINK'

    chromedriver = '/Users/arystanbekabdrahmanov/PycharmProjects/parser_kaspi_shop/chromedriver'
    options = webdriver.ChromeOptions()
    options.add_argument('headless')  # для открытия headless-браузера
    browser = webdriver.Chrome(executable_path=chromedriver, options=options)

    url = ''
    if choose_category == 'Айфоны':
        url = 'https://kaspi.kz/shop/c/smartphones/?q=%3AproductClass%3AApple+iPhone&page='
    elif choose_category == 'Шины':
        url = 'https://kaspi.kz/shop/c/passenger%20car%20tires/' \
              '?q=%3Acategory%3APassenger+car+tires%3Acategory%3ATires&page='
    elif choose_category == 'Морозилки':
        url = 'https://kaspi.kz/shop/c/freezers/?page='
    elif choose_category == 'Наушники':
        url = 'https://kaspi.kz/shop/c/headphones/?page='
    elif choose_category == 'Ноутбуки офисные':
        url = 'https://kaspi.kz/shop/c/notebooks/?q=%3AproductClass%3AДля+работы+и+учёбы&page='
    elif choose_category == 'Ноутбуки игровые':
        url = 'https://kaspi.kz/shop/c/notebooks/?q=%3AproductClass%3AИгровые&page='
    elif choose_category == 'Кондиционеры':
        url = 'https://kaspi.kz/shop/c/air%20conditioners/?q=%3AproductClass%3AВ+комнату&page='
    elif choose_category == 'мониторы':
        url = 'https://kaspi.kz/shop/c/monitors/?page='
    else:
        print('Не тупи, вводи из предложенного)))')

    url2 = url
    next_page_check = False
    while not next_page_check:
        url = url2 + str(i)

        browser.get(url)
        html = browser.page_source

        soup = BeautifulSoup(html, 'lxml')
        divs = soup.find(
            'div',
            class_="item-cards-grid__cards").find_all('div', class_="item-card ddl_product ddl_product_link undefined")

        try:
            browser.get(url)
            html = browser.page_source

            soup = BeautifulSoup(html, 'lxml')
            divs = soup.find(
                'div',
                class_="item-cards-grid__cards").find_all('div', class_="item-card ddl_product ddl_product_link undefined")

        except:
            print('--------- next page not found')
            break

        for d in divs:
            price = d.find('span', class_="item-card__prices-price").text.replace(' ', '').replace('₸', '')
            name = d.find('a', class_="item-card__name-link").text
            product_id = d.get('data-product-id')
            link = d.find('a', class_='item-card__name-link').get('href')

            # отдельный реквест для того, чтобы перейти на страницу продукта #####
            time.sleep(1)

            r1 = browser.get(link)
            product_html = browser.page_source

            soup_info = BeautifulSoup(product_html, 'lxml')
            try:
                details = soup_info('div', class_='item__description-text')
                details = details[0].text

            except:
                details = 'no descriptions'

            j += 1
            cell = sheet.cell(row=j, column=1)
            cell.value = j - 1
            cell = sheet.cell(row=j, column=2)
            cell.value = product_id
            cell = sheet.cell(row=j, column=3)
            cell.value = name
            cell = sheet.cell(row=j, column=4)
            cell.value = price
            cell = sheet.cell(row=j, column=5)
            cell.value = details
            cell = sheet.cell(row=j, column=6)
            cell.value = link
            print(j - 1, name,'|', price, 'тенге |||| ===========')
            file_name = 'prices_from_kaspi_kz.xlsx'
            wb.save(file_name)
        if i % 2 == 0:
            time.sleep(1)
        next_page_check = soup.findAll('li', class_="pagination__el _disabled")
        if next_page_check and next_page_check[0].contents[0] == 'Следующая →':
            break
        else:
            i += 1
            next_page_check = False
    send_mail(mail_mysql)



bot.polling(none_stop=True, interval=0)


